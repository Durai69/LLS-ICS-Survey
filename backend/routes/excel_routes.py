from flask import Blueprint, request, send_file, jsonify, g
from sqlalchemy.orm import Session
import pandas as pd
import io
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import defaultdict
from ..models import SurveyResponse, Department, User, Question, SurveySubmission, Answer
from ..database import SessionLocal
from backend.utils.paseto_utils import paseto_required, get_paseto_identity

excel_bp = Blueprint('excel', __name__)

def filter_responses_by_time_period(responses, time_period):
    if not time_period:
        return responses
    filtered = []
    try:
        parts = time_period.split()
        years = parts[0].split('-')
        start_year = int(years[0])
        end_year = int(years[1])
        survey_part = parts[1] if len(parts) > 1 else None

        for resp in responses:
            if not resp.submitted_at:
                continue
            year = resp.submitted_at.year
            if year < start_year or year > end_year:
                continue
            if survey_part:
                month = resp.submitted_at.month
                if survey_part == '1st' and month > 6:
                    continue
                if survey_part == '2nd' and month <= 6:
                    continue
            filtered.append(resp)
    except Exception:
        filtered = responses
    return filtered

@excel_bp.route('/api/export', methods=['GET'])
@paseto_required()
def export_excel():
    export_type = request.args.get('type')
    time_period = request.args.get('timePeriod')

    db: Session = SessionLocal()
    try:
        current_username = get_paseto_identity()
        user = db.query(User).filter(User.username == current_username).first()
        if not user:
            return jsonify({"error": "User not found"}), 404

        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')

        import logging
        logging.basicConfig(level=logging.INFO)
        logger = logging.getLogger("excel_export")

        logger.info(f"Export type: {export_type}, Time period: {time_period}")

        if export_type == 'My Submitted Surveys':
            # Get all submissions made by the user with status 'Submitted'
            submissions = db.query(SurveySubmission).filter(
                SurveySubmission.submitter_user_id == user.id,
                SurveySubmission.status == 'Submitted'
            ).all()
            
            df_data = []
            for idx, submission in enumerate(submissions, 1):
                dept = db.query(Department).filter(Department.id == submission.rated_department_id).first()
                for answer in submission.answers:
                    question = db.query(Question).filter(Question.id == answer.question_id).first()
                    # Default remark from submission
                    remark = submission.rating_description if submission.rating_description else ""
                    # If rating is 1 or 2, try to get remark from SurveyResponse
                    if answer.rating_value in [1, 2]:
                        sr = db.query(SurveyResponse).filter(
                            SurveyResponse.survey_submission_id == submission.id,
                            SurveyResponse.question_id == answer.question_id,
                            SurveyResponse.rating == answer.rating_value
                        ).first()
                        if sr and sr.remark:
                            remark = sr.remark
                    df_data.append({
                        "SL No": idx,
                        "Date of Submission": submission.submitted_at.strftime('%d.%m.%Y') if submission.submitted_at else "",
                        "Department": dept.name if dept else "",
                        "Category": question.category if question and question.category else "General",
                        "Question": question.text if question else "",
                        "Rating": answer.rating_value if answer.rating_value is not None else "",
                        "Remark": remark,
                        "Suggestions": submission.suggestions if submission.suggestions else "",
                    })
            df = pd.DataFrame(df_data)
            
            # Remove empty rows from DataFrame before writing to Excel
            if not df.empty:
                df = df.dropna(how='all')
                df.reset_index(drop=True, inplace=True)
                
            # Write DataFrame to a temporary sheet, then load to manipulate
            # We will start writing data from row 3 (0-indexed startrow=2)
            # Then insert rows above it, shifting data down.
            # IMPORTANT: The startrow for df.to_excel should now be 0, as we insert rows *before* writing the DF
            df.to_excel(writer, sheet_name='My Submitted Surveys', index=False, startrow=0)
            filename_base = "my_submitted_surveys"
            
            writer.close()
            output.seek(0)
            
            wb = openpyxl.load_workbook(output)
            ws = wb['My Submitted Surveys']
            
            # Define common styles
            header_fill = PatternFill(start_color="00CCFFCC", end_color="00CCFFCC", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Insert 2 rows at the top for title and user info
            ws.insert_rows(1, 2)

            # A1:H1 - Main Title
            ws.merge_cells('A1:H1')
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = "My Submitted Surveys - Internal Customer Focus"
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.font = Font(bold=True, size=14)
            title_cell.fill = header_fill # Apply background to title
            ws.row_dimensions[1].height = 40 # Make the merged title row more spacious

            # Apply background color to entire row 2 first
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=2, column=col_idx).fill = header_fill

            # A2:C2 - User Info
            ws.merge_cells('A2:C2')
            user_info_cell = ws.cell(row=2, column=1)
            user_info_cell.value = f"User: {user.name if user.name else user.username}"
            user_info_cell.font = Font(bold=True)
            user_info_cell.alignment = Alignment(vertical='center')
            # Fill already applied to the whole row, but ensure merged cell explicitly has it
            user_info_cell.fill = header_fill 

            # G2:H2 - Generated Date
            current_date = datetime.date.today().strftime('%d.%m.%Y')
            ws.merge_cells('G2:H2')
            date_cell = ws.cell(row=2, column=7)
            date_cell.value = f"Generated on: {current_date}"
            date_cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            date_cell.font = Font(bold=True)
            # Fill already applied to the whole row, but ensure merged cell explicitly has it
            date_cell.fill = header_fill 
            
            # Row 3 (Original header row after insertion)
            for cell in ws[3]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # There should be no empty row 4 now due to startrow=0 and insert_rows.
            # However, if any unexpected empty rows appear between header (row 3) and data (row 4 onwards),
            # this robust check can still be used.
            # Check for and remove empty row 4 specifically, if it exists and has a background
            def is_row_empty_and_filled(row_obj):
                is_empty = True
                has_fill = False
                for cell in row_obj:
                    if cell.value:
                        is_empty = False
                    if cell.fill and cell.fill.fill_type == 'solid':
                        has_fill = True
                return is_empty and has_fill

            if ws.max_row >= 4: # Ensure row 4 exists
                row4_cells = list(ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=ws.max_column))[0]
                if is_row_empty_and_filled(row4_cells):
                    ws.delete_rows(4)
            
            # Apply borders and alignment to all cells
            # Data starts from row 4 (Excel row, which was original row 2 for dataframe)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
                for cell in row:
                    cell.border = thin_border
                    # Apply specific alignment to data cells (rows after header)
                    if cell.row > 3: # Data rows start from Excel row 4
                        if cell.column in [1, 2, 6]:  # SL No, Date, and Rating columns
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        else: # Text columns
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
            
            # Set column widths
            column_widths = {
                'A': 10,   # SL No
                'B': 18,   # Date of Submission
                'C': 25,   # Department
                'D': 22,   # Category
                'E': 40,   # Question
                'F': 12,   # Rating
                'G': 35,   # Remark
                'H': 35,   # Suggestions
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            output2 = io.BytesIO()
            wb.save(output2)
            output2.seek(0)
            
            current_date_str = datetime.date.today().strftime('%Y%m%d')
            final_filename = f"{filename_base}_{current_date_str}.xlsx"
            
            return send_file(output2,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True,
                             download_name=final_filename)

        elif export_type == 'My Action Plan':
            # Get all responses where the user's department is the to_department (receiving department)
            user_dept_id = user.department_id
            if not user_dept_id:
                return jsonify({"error": "User has no associated department"}), 400
                
            responses = db.query(SurveyResponse).filter(
                SurveyResponse.to_department_id == user_dept_id,
                SurveyResponse.overall_rating == None # <-- Added this filter
            ).all()
            logger.info(f"Total action plans before filtering: {len(responses)}")
            responses = filter_responses_by_time_period(responses, time_period)
            logger.info(f"Total action plans after filtering: {len(responses)}")
            
            # Get user's department name
            user_dept = db.query(Department).filter(Department.id == user_dept_id).first()
            user_dept_name = user_dept.name if user_dept else "Unknown Department"
            
            # Create Excel file with custom formatting
            df_data = []
            for idx, resp in enumerate(responses, 1):
                from_dept = db.query(Department).filter(Department.id == resp.from_department_id).first()
                acknowledged_str = "Acknowledged" if resp.acknowledged else "Not Acknowledged"
                
                df_data.append({
                    "SL No": idx,
                    "Date of Survey": resp.submitted_at.strftime('%d.%m.%Y') if resp.submitted_at else "",
                    "Department": from_dept.name if from_dept else "",
                    "Problem / Suggestion for Improvement": resp.explanation if resp.explanation else "",
                    "Action Planned": resp.action_plan if resp.action_plan else "",
                    "Responsibility": resp.responsible_person if resp.responsible_person else "",
                    "Target Date": resp.target_date.strftime('%d.%m.%Y') if resp.target_date else "",
                    "Status": acknowledged_str,
                })
            
            df = pd.DataFrame(df_data)
            
            # Remove empty rows from DataFrame before writing to Excel
            if not df.empty:
                df = df.dropna(how='all')
                df.reset_index(drop=True, inplace=True)

            df.to_excel(writer, sheet_name='My Action Plan', index=False, startrow=0)
            filename_base = "my_action_plan"
            
            writer.close()
            output.seek(0)
            
            wb = openpyxl.load_workbook(output)
            ws = wb['My Action Plan']
            
            # Define common styles
            header_fill = PatternFill(start_color="00CCFFCC", end_color="00CCFFCC", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Insert 2 rows at the top for title and department info
            ws.insert_rows(1, 2)

            # A1:H1 - Main Title
            ws.merge_cells('A1:H1')
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = "Activity Plan for Internal Customer Focus - Suggestion for improvement"
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.font = Font(bold=True, size=14)
            title_cell.fill = header_fill # Apply background to title
            ws.row_dimensions[1].height = 40 # Make the merged title row more spacious

            # Apply background color to entire row 2 first
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=2, column=col_idx).fill = header_fill

            # A2:B2 - Department Info
            ws.merge_cells('A2:B2')
            dept_info_cell = ws.cell(row=2, column=1)
            dept_info_cell.value = f"Department: {user_dept_name}"
            dept_info_cell.font = Font(bold=True)
            dept_info_cell.alignment = Alignment(vertical='center')
            dept_info_cell.fill = header_fill # Ensure merged cell explicitly has it

            # G2:H2 - Updated Date
            current_date = datetime.date.today().strftime('%d.%m.%Y')
            ws.merge_cells('G2:H2')
            date_cell = ws.cell(row=2, column=7)
            date_cell.value = f"Updated as on Date: {current_date}"
            date_cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            date_cell.font = Font(bold=True)
            date_cell.fill = header_fill # Ensure merged cell explicitly has it
                
            # Row 3 (Original header row after insertion)
            for cell in ws[3]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Apply borders and alignment to all cells
            # Data starts from row 4 (Excel row, which was original row 2 for dataframe)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
                for cell in row:
                    cell.border = thin_border
                    
                    # Apply specific alignment to data cells (rows after header)
                    if cell.row > 3: # Data rows start from Excel row 4
                        if cell.column in [1, 2, 7, 8]:  # SL No, Date, Target Date and Status columns
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        else: # Text columns
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
            
            # Set column widths
            column_widths = {
                'A': 8,  # SL No
                'B': 15,  # Date of Survey
                'C': 20,  # Department
                'D': 30,  # Problem/Suggestion
                'E': 30,  # Action Planned
                'F': 20,  # Responsibility
                'G': 15,  # Target Date
                'H': 15,  # Status
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Save the modified workbook
            output2 = io.BytesIO()
            wb.save(output2)
            output2.seek(0)
            
            current_date_str = datetime.date.today().strftime('%Y%m%d')
            final_filename = f"{filename_base}_{current_date_str}.xlsx"
            
            return send_file(output2,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True,
                             download_name=final_filename)

        elif export_type == 'My Overall Ratings':
            # 1. Get current user's department
            user_dept_id = user.department_id
            if not user_dept_id:
                return jsonify({"error": "User has no associated department"}), 400

            # 2. Get all departments (for columns)
            # Filter out the user's own department from the list of all departments
            all_departments = db.query(Department).filter(Department.id != user_dept_id).all()
            # Ensure the user's department is not in the list of departments for columns
            all_departments_for_columns = [dept for dept in all_departments if dept.id != user_dept_id]

            dept_id_to_name = {dept.id: dept.name for dept in all_departments_for_columns}

            # 3. Get all unique questions by (category, order, text)
            questions = db.query(Question).order_by(Question.category, Question.order).all()
            unique_questions = []
            seen = set()
            for q in questions:
                key = (q.category, q.order, q.text)
                if key not in seen:
                    unique_questions.append(q)
                    seen.add(key)
            # Group by category
            criteria_map = defaultdict(list)
            for q in unique_questions:
                criteria_map[q.category].append(q)

            # 4. Get all answers received by the user's department excluding self-rating
            answers = db.query(Answer).join(SurveySubmission).filter(
                SurveySubmission.rated_department_id == user_dept_id,
                SurveySubmission.submitter_department_id != user_dept_id
            ).all()

            # Build a nested dict: {criteria: {sub_criteria: {from_dept: [ratings]}}}
            data = {}
            for crit, qs in criteria_map.items():
                data[crit] = {}
                for q in qs:
                    # Initialize with only departments that are not the user's own
                    data[crit][q.text] = {dept.name: [] for dept in all_departments_for_columns}

            # Map answers by question and from department
            for answer in answers:
                if not answer.question_id or answer.rating_value is None:
                    continue
                q = db.query(Question).filter(Question.id == answer.question_id).first()
                if not q or not q.category:
                    continue
                # Get from_dept from SurveySubmission submitter_department_id
                submission = db.query(SurveySubmission).filter(SurveySubmission.id == answer.submission_id).first()
                if not submission:
                    continue
                from_dept = dept_id_to_name.get(submission.submitter_department_id, None)
                if from_dept and q.text in data[q.category]:
                    data[q.category][q.text][from_dept].append(answer.rating_value)

            # 5. Prepare Excel data
            excel_rows = []
            summary_by_criteria = {crit: {dept.name: 0 for dept in all_departments_for_columns} for crit in criteria_map}
            total_by_dept = {dept.name: 0 for dept in all_departments_for_columns}

            # Assuming 20 questions total (4 per category) with max rating 4 => max total score = 80
            max_total_score = 0
            for qs in criteria_map.values():
                max_total_score += len(qs) * 4

            ordered_criteria = ["QUALITY", "DELIVERY", "COMMUNICATION", "RESPONSIVENESS", "IMPROVEMENT"]

            main_sl_no = 1
            for crit in ordered_criteria:
                if crit not in criteria_map:
                    continue

                # Sort questions within the category by their order attribute
                sorted_questions = sorted(criteria_map[crit], key=lambda x: x.order)

                for idx, q in enumerate(sorted_questions, 1):
                    # Use a consistent sub-lettering, like (a), (b), (c)...
                    sub_letter = chr(96 + idx)
                    row = [main_sl_no, f"{crit}", f"({sub_letter}) {q.text}"]
                    for dept in all_departments_for_columns: # Iterate over filtered departments
                        ratings = data[crit][q.text][dept.name]
                        val = sum(ratings) if ratings else 0
                        row.append(val)
                        summary_by_criteria[crit][dept.name] += val
                        total_by_dept[dept.name] += val
                    excel_rows.append(row)
                    main_sl_no += 1

                # Add summary row for this criteria
                row = ["", "", f"Sum of {crit[0]}"]
                for dept in all_departments_for_columns: # Iterate over filtered departments
                    row.append(summary_by_criteria[crit][dept.name])
                excel_rows.append(row)

            # Add total and percentage rows
            total_row = ["", "", "Total (Sum of Q,D,C,R & I)"]
            percent_row = ["", "", "Total in % = ((Sum of Q,D,C,R & I)/80)*100"]

            # Use fixed 80 for percentage calculation as per source image formula
            fixed_max_score = 80

            for dept in all_departments_for_columns: # Iterate over filtered departments
                total = total_by_dept[dept.name]
                total_row.append(total)
                percent = round((total / fixed_max_score) * 100, 2) if fixed_max_score else 0
                percent_row.append(percent)

            excel_rows.append(total_row)
            excel_rows.append(percent_row)

            # 6. Write to Excel
            df = pd.DataFrame(excel_rows, columns=["Sl. No", "Category", "Evaluation Criteria"] + [dept.name for dept in all_departments_for_columns])
            df.to_excel(writer, sheet_name='My Overall Ratings', index=False, startrow=1)
            filename_base = "my_overall_ratings"

            writer.close()
            output.seek(0)

            # 7. Apply formatting using openpyxl
            wb = openpyxl.load_workbook(output)
            ws = wb['My Overall Ratings']

            # --- Start Formatting ---

            # General Styles
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
            rotated_align = Alignment(horizontal='center', vertical='center', text_rotation=90)
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Light blue fill

            # Format Title (Row 1)
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = "Internal Customer Satisfaction Survey Report - My Overall Ratings"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = center_align
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.alignment = center_align
            ws.row_dimensions[1].height = 30  # Add more space above and below title

            # Format Header Row (Row 2)
            # Merge A and B columns for header row
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
            header_cell = ws.cell(row=2, column=1)
            header_cell.value = "Sl. No"
            header_cell.font = bold_font
            header_cell.alignment = center_align
            header_cell.fill = header_fill

            # Set header for Category column (now column 3)
            category_cell = ws.cell(row=2, column=3)
            category_cell.value = "Evaluation Criteria"
            category_cell.font = bold_font
            category_cell.alignment = center_align
            category_cell.fill = header_fill

            # Clear the original header cell in column B (now merged)
            # Instead of setting value to None (which causes error on merged cell), delete the cell
            ws._cells.pop((2, 2), None)

            # Add background color and more space for row 2 (header)
            for cell in ws[2]:
                cell.fill = header_fill
            ws.row_dimensions[2].height = 25

            # Format Data Area (Row 3 onwards)
            for row_idx in range(3, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 35 # Set uniform row height
                
                is_summary_row = False
                eval_criteria_cell = ws.cell(row=row_idx, column=3)
                if eval_criteria_cell.value and (str(eval_criteria_cell.value).startswith("Sum of") or str(eval_criteria_cell.value).startswith("Total")):
                    is_summary_row = True

                if is_summary_row:
                    for cell in ws[row_idx]:
                        cell.fill = yellow_fill
                        cell.font = bold_font
                        cell.alignment = center_align
                else:
                    ws.cell(row=row_idx, column=1).alignment = center_align
                    ws.cell(row=row_idx, column=3).alignment = left_align
                    for col_idx in range(4, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).alignment = center_align

            # Merge Category Cells Vertically
            row_offset = 2
            current_row = row_offset + 1
            for crit in ordered_criteria:
                if crit not in criteria_map:
                    continue
                
                sorted_questions = sorted(criteria_map[crit], key=lambda x: x.order)
                n = len(sorted_questions)
                if n > 0:
                    merge_start_row = current_row
                    merge_end_row = current_row + n - 1
                    ws.merge_cells(start_row=merge_start_row, start_column=2, end_row=merge_end_row, end_column=2)
                    merged_cell = ws.cell(row=merge_start_row, column=2)
                    merged_cell.alignment = rotated_align
                    merged_cell.font = bold_font
                current_row += n + 1

            # Set Column Widths
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 5
            ws.column_dimensions['C'].width = 50
            # Adjust column width loop to account for filtered departments
            for i in range(4, 4 + len(all_departments_for_columns)): # Use all_departments_for_columns here
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12

            # Apply borders to all cells in the used range
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
            
            # --- End Formatting ---
            
            output2 = io.BytesIO()
            wb.save(output2)
            output2.seek(0)
            current_date_str = datetime.date.today().strftime('%Y%m%d')
            final_filename = f"{filename_base}_{current_date_str}.xlsx"
            return send_file(output2,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True,
                             download_name=final_filename)

        else:
            writer.close()
            return jsonify({"error": "Unknown export type"}), 400
    finally:
        db.close()

@excel_bp.route('/api/admin/reports', methods=['GET'])
def get_admin_reports():
    from_dept = request.args.get('fromDept')
    to_dept = request.args.get('toDept')
    time_period = request.args.get('timePeriod')

    db: Session = SessionLocal()
    query = db.query(SurveyResponse)
    if from_dept:
        from_dept_obj = db.query(Department).filter(Department.name == from_dept).first()
        if from_dept_obj:
            query = query.filter(SurveyResponse.from_department_id == from_dept_obj.id)
    if to_dept:
        to_dept_obj = db.query(Department).filter(Department.name == to_dept).first()
        if to_dept_obj:
            query = query.filter(SurveyResponse.to_department_id == to_dept_obj.id)
    responses = query.all()
    responses = filter_responses_by_time_period(responses, time_period)
    result = []
    for resp in responses:
        from_dept_obj = db.query(Department).filter(Department.id == resp.from_department_id).first()
        to_dept_obj = db.query(Department).filter(Department.id == resp.to_department_id).first()
        # Calculate avgRating if you want (for now, use resp.rating)
        result.append({
            "id": resp.id,
            "from_department": from_dept_obj.name if from_dept_obj else "",
            "to_department": to_dept_obj.name if to_dept_obj else "",
            "date": resp.submitted_at.strftime('%Y-%m-%d') if resp.submitted_at else "",
            "remark": resp.remark if resp.remark else "",
        })
    return jsonify(result)

@excel_bp.route('/api/admin/reports/export', methods=['GET'])
def export_admin_reports_excel():
    from_dept = request.args.get('fromDept')
    to_dept = request.args.get('toDept')
    time_period = request.args.get('timePeriod')

    db: Session = SessionLocal()
    query = db.query(SurveyResponse)
    if from_dept:
        from_dept_obj = db.query(Department).filter(Department.name == from_dept).first()
        if from_dept_obj:
            query = query.filter(SurveyResponse.from_department_id == from_dept_obj.id)
    if to_dept:
        to_dept_obj = db.query(Department).filter(Department.name == to_dept).first()
        if to_dept_obj:
            query = query.filter(SurveyResponse.to_department_id == to_dept_obj.id)
    responses = query.all()
    responses = filter_responses_by_time_period(responses, time_period)

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Main sheet
    main_data = []
    for resp in responses:
        from_dept_obj = db.query(Department).filter(Department.id == resp.from_department_id).first()
        to_dept_obj = db.query(Department).filter(Department.id == resp.to_department_id).first()
        main_data.append({
            "Date": resp.submitted_at.strftime('%d-%m-%Y') if resp.submitted_at else "",
            "From Department": from_dept_obj.name if from_dept_obj else "",
            "To Department": to_dept_obj.name if to_dept_obj else "",
            "Overall Rating": resp.overall_rating if resp.overall_rating is not None else "",
        })
    main_df = pd.DataFrame(main_data)
    main_df.to_excel(writer, sheet_name='Survey Reports', index=False)

    # Action Plan sheet
    action_data = []
    for resp in responses:
        from_dept_obj = db.query(Department).filter(Department.id == resp.from_department_id).first()
        to_dept_obj = db.query(Department).filter(Department.id == resp.to_department_id).first()
        category = ""
        if resp.question_id:
            question = db.query(Question).filter(Question.id == resp.question_id).first()
            category = question.category if question and question.category else ""
        acknowledged_str = "Acknowledged" if resp.acknowledged else "Not Acknowledged"
        action_data.append({
            "Date": resp.submitted_at.strftime('%d-%m-%Y') if resp.submitted_at else "",
            "From Department": from_dept_obj.name if from_dept_obj else "",
            "To Department": to_dept_obj.name if to_dept_obj else "",
            "Rating Value": resp.rating if resp.rating is not None else "",
            "Category": category,
            "Explanation": resp.explanation if resp.explanation else "",
            "Action Plan": resp.action_plan if resp.action_plan else "",
            "Responsible Person": resp.responsible_person if resp.responsible_person else "",
            "Target Date": resp.target_date.strftime('%d-%m-%Y') if resp.target_date else "",
            "Acknowledged": acknowledged_str,
        })
    action_df = pd.DataFrame(action_data)
    action_df.to_excel(writer, sheet_name='Action Plans', index=False)

    writer.close()
    output.seek(0)

    # Set column width to 20 for all columns in all sheets
    wb = openpyxl.load_workbook(output)
    for ws in wb.worksheets:
        for col in ws.columns:
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = 20

    output2 = io.BytesIO()
    wb.save(output2)
    output2.seek(0)

    current_date_str = datetime.date.today().strftime('%Y%m%d')
    final_filename = f"admin_survey_reports_{current_date_str}.xlsx"
    return send_file(output2,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=final_filename)